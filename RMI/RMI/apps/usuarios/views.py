#-*- coding: utf-8 -*-
from django.shortcuts import render
from django.shortcuts import render_to_response
from django.template import RequestContext
from RMI.apps.usuarios.forms import *
from django.contrib.auth import login, logout, authenticate
from django.http import HttpResponseRedirect
from django.contrib.auth.models import User ##
from RMI.apps.usuarios.models import *
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, InvalidPage#paginator
from django.http.response import HttpResponse
from django.core.mail import EmailMultiAlternatives
from django.contrib import messages
from openpyxl import Workbook
from django.conf import settings
from django.core.files import File
import os 


# Create your views here.
def login_view(request):
	mensaje = ""
	if request.user.is_authenticated():#verificacmos si el usuario ya esta authenticado o logueado
		return HttpResponseRedirect('/index/')#si esta logueado lo redirigimos a la pagina principal
	else: #si no esta authenticado 
		if request.method == "POST":
			formulario = Login_form(request.POST) #creamos un objeto de Loguin_form
			if formulario.is_valid(): #si la informacion enviada es correcta		
				usu= formulario.cleaned_data['usuario'] #guarda informacion ingresada del formulario
				pas= formulario.cleaned_data['clave'] #guarda informacion ingresada del formulario
				usuario = authenticate(username = usu,password = pas)#asigna la autenticacion del usuario
				if usuario is not None and usuario.is_active:#si el usuario no es nulo y esta activo
					login(request,usuario)#se loguea al sistema con la informacion de usuario
					return HttpResponseRedirect('/index/')#redirigimos a la pagina principal
				else:
					mensaje = "usuario y/o clave incorrecta"
		formulario = Login_form() #creamos un formulario nuevo limpio
		ctx = {'form':formulario, 'mensaje':mensaje}#variable de contexto para pasar info a login.html
		return render_to_response('usuarios/login.html',ctx, context_instance = RequestContext(request))

def logout_view(request):
	logout(request)# funcion de django importda anteriormente
	return HttpResponseRedirect('/')# redirigimos a la pagina principal

def user_view(request): 
	us = User.objects.get(id= request.user.id)
	ctx={'user':us}
	return render_to_response('usuarios/user.html',ctx,context_instance = RequestContext(request))

def admin_user_view(request, id_user): 
	us = User.objects.get(id= id_user)
	ctx={'usuario':us}
	return render(request,'usuarios/admin_user.html',ctx)			

def edit_user_view(request):
	info = ""	
	info_enviado = False
	us = User.objects.get(id = request.user.id)
	user = User_profile.objects.get(user=us)
	#formulario = UserForm()
	if request.method == "POST":
		formulario = UserForm(request.POST, request.FILES, instance = us)
		form_user= User_profile_Form(request.POST, request.FILES, instance = user)		
		if formulario.is_valid():
			info_enviado = True
			#telefono = formulario.cleaned_data['telefono']
			#clave = formulario.cleaned_data['password']
			#formulario.password = us.set_password(clave)
			edit_user = formulario.save(commit = False)
			edit_user_telefono = form_user.save(commit = False)
			#edit_user_telefono.telefono= telefono
			#formulario.save_m2m()
			#edit_user.status = True
			edit_user_telefono.save()
			edit_user.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/user/')
	else:
		formulario = UserForm(instance = us)
		form_user = User_profile_Form(instance = user)
	ctx = {'form':formulario, 'form_user':form_user,'informacion':info, 'info_enviado':info_enviado}	
	return render_to_response('usuarios/edit_user.html',ctx,context_instance = RequestContext(request))

def admin_edit_user_view(request,id_user):
	info = ""	
	info_enviado = False
	us = User.objects.get(id = id_user)
	user = User_profile.objects.get(user=us)
	if request.method == "POST":
		formulario = AdminUserForm(request.POST, request.FILES, instance = us)	
		form_user= User_profile_Form(request.POST, request.FILES, instance = user)		
		if formulario.is_valid():
			info_enviado = True
			#clave = formulario.cleaned_data['password']
			#formulario.password = us.set_password(clave)
			edit_user = formulario.save(commit = False)	
			edit_user_telefono = form_user.save(commit = False)		
			#edit_user.status = True
			edit_user_telefono.save()
			edit_user.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/instructores/')
	else:
		formulario = AdminUserForm(instance = us)
		form_user = User_profile_Form(instance = user)
	ctx = {'form':formulario, 'form_user':form_user,'informacion':info, 'info_enviado':info_enviado,'usuario':us}	
	return render_to_response('usuarios/admin_edit_user.html',ctx,context_instance = RequestContext(request))		

def edit_password_view(request):
	info = ""	
	info_enviado = False
	us = User.objects.get(id = request.user.id)
	if request.method == "POST":
		formulario = PasswordForm(request.POST, request.FILES, instance = us)		
		if formulario.is_valid():
			info_enviado = True
			clave = formulario.cleaned_data['password']
			usu= us.username #guarda informacion ingresada del formulario
			pas= clave #guarda informacion ingresada del formulario
			formulario.password = us.set_password(clave)
			edit_user = formulario.save(commit = False)			
			#formulario.save_m2m()
			#edit_user.status = True
			edit_user.save()
			info = "Guardado Satisfactoriamente"
			usuario = authenticate(username = usu,password = pas)#asigna la autenticacion del usuario
			if usuario is not None and usuario.is_active:#si el usuario no es nulo y esta activo
				login(request,usuario)#se loguea al sistema con la informacion de usuario
				return HttpResponseRedirect('/password_guardado_user/')#redirigimos a la pagina principal
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/index/')
	else:
		formulario = PasswordForm(instance = us)
	ctx = {'form':formulario, 'informacion':info, 'info_enviado':info_enviado}	
	return render_to_response('usuarios/edit_password.html',ctx,context_instance = RequestContext(request))

def admin_edit_password_view(request,id_user):
	info = ""	
	info_enviado = False
	us = User.objects.get(id = id_user)
	if request.method == "POST":
		formulario = PasswordForm(request.POST, request.FILES, instance = us)		
		if formulario.is_valid():
			info_enviado = True
			clave = formulario.cleaned_data['password']
			formulario.password = us.set_password(clave)
			edit_user = formulario.save(commit = False)
			#formulario.save_m2m()
			#edit_user.status = True
			edit_user.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			return HttpResponseRedirect ('/password_guardado_admin/')
	else:
		formulario = PasswordForm(instance = us)
	ctx = {'form':formulario, 'informacion':info, 'info_enviado':info_enviado,'user':us}	
	return render_to_response('usuarios/admin_edit_password.html',ctx,context_instance = RequestContext(request))				


def index_view(request):
	usuario = User.objects.get(id= request.user.id)
	mes_enero_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='ENERO')
	mes_febrero_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='FEBRERO')
	mes_marzo_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='MARZO')
	mes_mayo_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='MAYO')
	mes_abril_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='ABRIL')
	mes_junio_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='JUNIO')
	mes_julio_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='JULIO')
	mes_agosto_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='AGOSTO')
	mes_septiembre_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='SEPTIEMBRE')
	mes_octubre_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='OCTUBRE')
	mes_noviembre_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='NOVIEMBRE')
	mes_diciembre_instructor = Reporte_Mensual.objects.filter(usuario= usuario,mes='DICIEMBRE')

	mes_enero_admin = Reporte_Mensual.objects.filter(mes='ENERO')
	mes_febrero_admin = Reporte_Mensual.objects.filter(mes='FEBRERO')
	mes_marzo_admin = Reporte_Mensual.objects.filter(mes='MARZO')
	mes_mayo_admin = Reporte_Mensual.objects.filter(mes='MAYO')
	mes_abril_admin = Reporte_Mensual.objects.filter(mes='ABRIL')
	mes_junio_admin = Reporte_Mensual.objects.filter(mes='JUNIO')
	mes_julio_admin = Reporte_Mensual.objects.filter(mes='JULIO')
	mes_agosto_admin = Reporte_Mensual.objects.filter(mes='AGOSTO')
	mes_septiembre_admin = Reporte_Mensual.objects.filter(mes='SEPTIEMBRE')
	mes_octubre_admin = Reporte_Mensual.objects.filter(mes='OCTUBRE')
	mes_noviembre_admin = Reporte_Mensual.objects.filter(mes='NOVIEMBRE')
	mes_diciembre_admin = Reporte_Mensual.objects.filter(mes='DICIEMBRE')

	ctx={'usuario':usuario,
		'mes_enero_instructor':mes_enero_instructor,
		'mes_febrero_instructor':mes_febrero_instructor,
		'mes_marzo_instructor':mes_marzo_instructor,
		'mes_mayo_instructor':mes_mayo_instructor,
		'mes_abril_instructor':mes_abril_instructor,
		'mes_junio_instructor':mes_junio_instructor,
		'mes_julio_instructor':mes_julio_instructor,
		'mes_agosto_instructor':mes_agosto_instructor,
		'mes_septiembre_instructor':mes_septiembre_instructor,
		'mes_octubre_instructor':mes_octubre_instructor,
		'mes_noviembre_instructor':mes_noviembre_instructor,
		'mes_diciembre_instructor':mes_diciembre_instructor,
		'mes_enero_admin':mes_enero_admin,
		'mes_febrero_admin':mes_febrero_admin,
		'mes_marzo_admin':mes_marzo_admin,
		'mes_mayo_admin':mes_mayo_admin,
		'mes_abril_admin':mes_abril_admin,
		'mes_junio_admin':mes_junio_admin,
		'mes_julio_admin':mes_julio_admin,
		'mes_agosto_admin':mes_agosto_admin,
		'mes_septiembre_admin':mes_septiembre_admin,
		'mes_octubre_admin':mes_octubre_admin,
		'mes_noviembre_admin':mes_noviembre_admin,
		'mes_diciembre_admin':mes_diciembre_admin
		}
	return render(request, 'usuarios/index.html',ctx) 

def password_guardado_user_view(request):
	
	return render(request, 'usuarios/password_guardado_user.html') 	

def password_guardado_admin_view(request):
	
	return render(request, 'usuarios/password_guardado_admin.html') 		 


def eliminar_instructor_view(request,id_user):	
	usuario = User.objects.get(id= id_user)
	ctx={'usuario':usuario}
	return render(request, 'usuarios/eliminar_instructor.html',ctx)

def confirmar_eliminar_instructor_view(request,id_user):	
	
	info = "inicializando"
	try:
		usuario = User.objects.get(id= id_user)
		user_telefono = User_profile.objects.get(user= usuario)
		reporte = Reporte_Mensual.objects.filter(usuario= usuario)
		usuario.delete()			
		user_telefono.delete()
		if reporte:
			for r in reporte:
				r.delete()
		#rep.adjunto_exel.url.delete()
		info = "El instructor ha sido eliminado correctamente"
		return HttpResponseRedirect('/instructores/page/1')
	except:
		info = "EL reporte no se puede eliminar"
		#return render_to_response('home/productos.html', context_instance = RequestContext(request))
		return HttpResponseRedirect('/index/')		

def register_view(request):
	form = RegisterForm()
	user_form = User_profile_Form()
	if request.method == "POST":
		form = RegisterForm(request.POST)
		user_form = User_profile_Form(request.POST)
		if form.is_valid():
			nombres = form.cleaned_data['first_name']
			apellidos = form.cleaned_data['last_name']
			usuario = form.cleaned_data['username']
			email = form.cleaned_data['email']
			password_one = form.cleaned_data['password_one']
			password_two = form.cleaned_data['password_two']
			telefono = form.cleaned_data['telefono']
			supervisor  = form.cleaned_data['supervisor']
			u = User.objects.create_user(first_name=nombres,last_name=apellidos,username=usuario,email=email,password=password_one)
			user = user_form.save(commit=False)
			user.telefono= telefono
			user.supervisor = supervisor
			user.user=u
			#user= form.save(commit=False)
			#user.user_profile.telefono=telefono
			#user.save()
			#user = u.save(commit=False)
			#user.user_profile.telefono=telefono
			user.save()			
			u.save()# Guarda el objeto
			#u.user_profile.telefono=telefono
			#u.save()
			ctx = {'usuario':u}
			return render_to_response('usuarios/thanks_register.html',ctx,context_instance=RequestContext(request))
		else:		
			ctx = {'form':form}
			return render_to_response('usuarios/register.html',ctx,context_instance=RequestContext(request))
	ctx = {'form':form}
	return render_to_response('usuarios/register.html',ctx,context_instance=RequestContext(request))

def instructores_view(request,pagina):
	lista_instructores = User.objects.filter(is_staff=False)	
	primera = "<<Primera"
	ultima = "Ultima>>"	
	query = request.GET.get('q','')     
	if query:
		qset = (
			Q(first_name__icontains=query)|
			Q(last_name__icontains=query)|
			Q(username__icontains=query)|
			Q(email__icontains=query)
		)
		results = User.objects.filter(qset).distinct()  
		mostrar = False      
	else:
		mostrar = True
		results = []

	#lista_prod = Producto.objects.filter(status = True)#SELECT * from Producto WHERE status= True
	paginator = Paginator(lista_instructores, 15) 
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		instructores = paginator.page(page)
	except (EmptyPage,InvalidPage):
		instructores = paginator.page(paginator.num_pages)	


	return render_to_response('usuarios/instructores.html', {
		"results": results,
		"query": query,
		"mostrar": mostrar,
		"instructores":instructores,
		"lista_instructores":lista_instructores, 
		"primera":primera,
		"ultima":ultima,       
	},context_instance=RequestContext(request))

	#ctx={'instructores':instructores}
	#return render(request, 'usuarios/instructores.html',ctx) 	
########################### SUBIR REPORTES POR MESES ##################################################
def reportes_view(request, id_mes):
	meses = { '1':'ENERO', '2':'FEBRERO', '3':'MARZO', '4':'ABRIL', '5':'MAYO','6':'JUNIO','7':'JULIO','8':'AGOSTO','9':'SEPTIEMBRE','10':'OCTUBRE','11':'NOVIEMBRE','12':'DICIEMBRE' }	
	mes = meses[id_mes]		
	no_tiene=False
	mes_enero=False
	mes_febrero=False
	mes_marzo=False
	mes_mayo=False
	mes_abril=False
	mes_junio=False
	mes_julio=False
	mes_agosto=False
	mes_septiembre=False
	mes_octubre=False
	mes_noviembre=False
	mes_diciembre=False

	if mes=='ENERO':
		mes_enero=True
	elif mes=='FEBRERO':
		mes_febrero=True
	elif mes=='MARZO':
		mes_marzo=True
	elif mes=='ABRIL':
		mes_abril=True
	elif mes=='MAYO':
		mes_mayo=True		
	elif mes=='JUNIO':
		mes_junio=True
	elif mes=='JULIO':
		mes_julio=True
	elif mes=='AGOSTO':
		mes_agosto=True
	elif mes=='SEPTIEMBRE':
		mes_septiembre=True
	elif mes=='OCTUBRE':
		mes_octubre=True
	elif mes=='NOVIEMBRE':
		mes_noviembre=True
	elif mes=='DICIEMBRE':
		mes_diciembre=True

	reporte_validar = Reporte_Mensual.objects.filter(mes= mes,usuario__id= request.user.id )
	user = User.objects.get(id= request.user.id )
	area = user.user_profile.area
	reportes = Reporte_Mensual.objects.filter(mes= mes,usuario__user_profile__area=area)	
	usuarios_area= User.objects.filter(user_profile__area=user.user_profile.area)	
	if user.user_profile.area=='No tiene':
		no_tiene=True
	#nombre_adjunto = ""	
	ctx={'reporte_validar':reporte_validar, 'mes':mes, 'user':user, 'usuarios_area':usuarios_area, 'no_tiene':no_tiene, 'reportes':reportes,
		'mes_enero':mes_enero,
		'mes_febrero':mes_febrero,
		'mes_marzo':mes_marzo,
		'mes_mayo':mes_mayo,
		'mes_abril':mes_abril,
		'mes_junio':mes_junio,
		'mes_julio':mes_julio,
		'mes_agosto':mes_agosto,
		'mes_septiembre':mes_septiembre,
		'mes_octubre':mes_octubre,
		'mes_noviembre':mes_noviembre,
		'mes_diciembre':mes_diciembre
		}

	return render(request, 'usuarios/reportes.html',ctx) 	

def reportes_lider_view(request, mes, id_user):
	meses = { 'ENERO': '1', 'FEBRERO': '2', 'MARZO': '3', 'ABRIL':'4', 'MAYO':'5', 'JUNIO':'6', 'JULIO':'7', 'AGOSTO':'8', 'SEPTIEMBRE':'9', 'OCTUBRE':'10', 'NOVIEMBRE':'11', 'DICIEMBRE':'12' }	
	id_mes = meses[mes]	
	
	reporte_validar = Reporte_Mensual.objects.filter(mes= mes,usuario__id= id_user )
	user = User.objects.get(id= id_user )	
	nombre_adjunto = ""	
	info_enviado = False
	if request.method == "POST":		
		form = Reporte(request.POST, request.FILES)
		if form.is_valid():
			info_enviado = True
			nombre_adjunto = form.cleaned_data['nombre_adjunto']											
			reporte = form.save(commit = False)
			reporte.usuario = user	
			reporte.mes = mes			
			reporte.save()

			if id_mes=='1':
				user.user_profile.enero=True
			elif id_mes=='2':
				user.user_profile.febrero=True
			elif id_mes=='3':
				user.user_profile.marzo=True
			elif id_mes=='4':
				user.user_profile.abril=True
			elif id_mes=='5':
				user.user_profile.mayo=True
			elif id_mes=='6':
				user.user_profile.junio=True
			elif id_mes=='7':
				user.user_profile.julio=True
			elif id_mes=='8':
				user.user_profile.agosto=True
			elif id_mes=='9':
				user.user_profile.septiembre=True
			elif id_mes=='10':
				user.user_profile.octubre=True
			elif id_mes=='11':
				user.user_profile.noviembre=True
			elif id_mes=='12':
				user.user_profile.diciembre=True

			user.user_profile.save()
			#to_admin = 'lgonzalez21@misena.edu.co'
			to_admin = 'fkenneth228@gmail.com'
			to_admin_2 = 'alexandrajr@misena.edu.co'
			to_admin_3 = 'mcperezp@sena.edu.co'
			#to_user = correo
			html_content_admin = "<p>El instructor <b>%s %s</b> ha subido un archivo</p><br><p>http://reportemensualinstructor.herokuapp.com%s</p>"%(user.first_name,user.last_name,reporte.adjunto_exel.url)
			html_content_admin2 = "<p>El instructor <b>%s %s</b> ha subido un archivo</p>"%(user.first_name,user.last_name)
			#html_content_user = "<p><b>Solicitud de servicio: </b>%s</p> <!--<p><b>Codigo de radicado:</b> %s</p>--> <p><b>Apreciado usuario, su solicitud será respondida en un termino maximo de 24 horas, por favor tenga en cuenta siguientes instrucciones:</b></p> 1. Debe imprimir únicamente copia que hace referencia al banco, importante: DEBE IMPRIMIR EL RECIBO EN IMPRESORA LASER.<br>2. Debe hacer consignación en sucursales Bancolombia(no corresponsales bancarios).<br>3. Una vez consigne o cancele su recibo debe hacerlos llegar a las oficinas de coordinación académica según su solicitud ,en este caso:  %s.<br>4. La consignación debe hacerse el mismo día que se genera el recibo."%(servicio,codigo_parsear,servicio_usuario)

			msg = EmailMultiAlternatives('Instructor %s %s, mes %s'%(user.first_name,user.last_name,mes), html_content_admin, 'from@gmail.com',[to_admin])
			msg2 = EmailMultiAlternatives('Instructor %s %s, mes %s'%(user.first_name,user.last_name,mes), html_content_admin2, 'from@gmail.com',[to_admin_2])
			msg3 = EmailMultiAlternatives('Instructor %s %s, mes %s'%(user.first_name,user.last_name,mes), html_content_admin2, 'from@gmail.com',[to_admin_3])
			#msg2 = EmailMultiAlternatives('Solicitud de recibo de consignacion %s (Recuerde esto al momento de obtener el recibo)'%(codigo_parsear), html_content_user, 'from@gmail.com',[to_user])
			file_path = settings.MEDIA_ROOT + str(reporte.adjunto_exel)
			#print(file_path)
			fd = open("%s"%(file_path),"rb")
			msg.attach('%s'%(reporte.adjunto_exel.name),fd.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			fd2 = open("%s"%(file_path),"rb")
			msg2.attach('%s'%(reporte.adjunto_exel.name),fd2.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			fd3 = open("%s"%(file_path),"rb")
			msg3.attach('%s'%(reporte.adjunto_exel.name),fd3.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			msg.attach_alternative(html_content_admin,'text/html')			
			msg2.attach_alternative(html_content_admin2,'text/html')	
			msg3.attach_alternative(html_content_admin2,'text/html')	
			#		
			msg.send()
			msg2.send()
			msg3.send()
			#form_status= formulario.save(commit=False)
			#if msg.send:
			#	form_status.status_admin=True
			#	info_enviado_admin = True
			#if msg2.send:
			#	form_status.status_user=True
			#	info_enviado_user = True

			#form_status.save()

	else:
		form = Reporte()

	ctx={'reporte_validar':reporte_validar, 'form':form, 'mes':mes, 'user':user, 'id_mes':id_mes}

	return render(request, 'usuarios/reportes_lider.html',ctx) 		
######################################################################################
'''
def enero_view(request):
	reporte_validar = Reporte_Mensual.objects.filter(mes= 'Enero',usuario__id= request.user.id )	
	nombre_adjunto = ""	
	info_enviado = False
	if request.method == "POST":		
		form = Reporte(request.POST, request.FILES)
		if form.is_valid():
			info_enviado = True
			nombre_adjunto = form.cleaned_data['nombre_adjunto']											
			reporte = form.save(commit = False)
			reporte.usuario = request.user	
			reporte.mes = 'Enero'					
			reporte.save()					
	else:
		form = Reporte()

	ctx={'reporte_validar':reporte_validar, 'form':form}

	return render(request, 'usuarios/enero.html',ctx) 	
'''
############################################### 	BORRAR REPORTES MESES    ############################################
def del_reporte_view(request, id_reporte, id_user):
	info = "inicializando"
	user = User.objects.get(id= id_user )	
	try:
		rep = Reporte_Mensual.objects.get(id = id_reporte)
		mes = rep.mes
		meses = { 'ENERO': '1', 'FEBRERO': '2', 'MARZO': '3', 'ABRIL':'4', 'MAYO':'5', 'JUNIO':'6', 'JULIO':'7', 'AGOSTO':'8', 'SEPTIEMBRE':'9', 'OCTUBRE':'10', 'NOVIEMBRE':'11', 'DICIEMBRE':'12' }	
		mes_num = meses[mes]
		id_mes=mes_num
		print mes	
		if id_mes=='1':
			user.user_profile.enero=False
		elif id_mes=='2':
			user.user_profile.febrero=False
		elif id_mes=='3':
			user.user_profile.marzo=False
		elif id_mes=='4':
			user.user_profile.abril=False
		elif id_mes=='5':
			user.user_profile.mayo=False
		elif id_mes=='6':
			user.user_profile.junio=False
		elif id_mes=='7':
			user.user_profile.julio=False
		elif id_mes=='8':
			user.user_profile.agosto=False
		elif id_mes=='9':
			user.user_profile.septiembre=False
		elif id_mes=='10':
			user.user_profile.octubre=False
		elif id_mes=='11':
			user.user_profile.noviembre=False
		elif id_mes=='12':
			user.user_profile.diciembre=False		

		rep.delete()
		user.user_profile.save()
		#rep.adjunto_exel.url.delete()
		info = "El reporte ha sido eliminado correctamente"
		return HttpResponseRedirect('/reportes_lider/%s/%s'%(mes,id_user))
	except:
		info = "EL reporte no se puede eliminar"
		#return render_to_response('home/productos.html', context_instance = RequestContext(request))
		return HttpResponseRedirect('/index/')	
'''
def del_reporte_enero_view(request, id_reporte):
	info = "inicializando"
	try:
		rep = Reporte_Mensual.objects.get(id = id_reporte)
		rep.delete()
		info = "El reporte ha sido eliminado correctamente"
		return HttpResponseRedirect('/enero/')
	except:
		info = "EL reporte no se puede eliminar"
		#return render_to_response('home/productos.html', context_instance = RequestContext(request))
		return HttpResponseRedirect('/index/')	

				

#############################################	CONSULTAR REPORTES POR MES #####################################################

'''
def reporte_exel_view(request,id_mes):
	meses = { '1':'ENERO', '2':'FEBRERO', '3':'MARZO', '4':'ABRIL', '5':'MAYO','6':'JUNIO','7':'JULIO','8':'AGOSTO','9':'SEPTIEMBRE','10':'OCTUBRE','11':'NOVIEMBRE','12':'DICIEMBRE' }	
	mes = meses[id_mes]	

	if id_mes=='1':
		lista_consultar = User.objects.filter(user_profile__enero=False,is_staff=False)
	elif id_mes=='2':
		lista_consultar = User.objects.filter(user_profile__febrero=False,is_staff=False)
	elif id_mes=='3':
		lista_consultar = User.objects.filter(user_profile__marzo=False,is_staff=False)
	elif id_mes=='4':
		lista_consultar = User.objects.filter(user_profile__abril=False,is_staff=False)
	elif id_mes=='5':
		lista_consultar = User.objects.filter(user_profile__mayo=False,is_staff=False)
	elif id_mes=='6':
		lista_consultar = User.objects.filter(user_profile__junio=False,is_staff=False)
	elif id_mes=='7':
		lista_consultar = User.objects.filter(user_profile__julio=False,is_staff=False)
	elif id_mes=='8':
		lista_consultar = User.objects.filter(user_profile__agosto=False,is_staff=False)
	elif id_mes=='9':
		lista_consultar = User.objects.filter(user_profile__septiembre=False,is_staff=False)
	elif id_mes=='10':
		lista_consultar = User.objects.filter(user_profile__octubre=False,is_staff=False)
	elif id_mes=='11':
		lista_consultar = User.objects.filter(user_profile__noviembre=False,is_staff=False)
	elif id_mes=='12':
		lista_consultar = User.objects.filter(user_profile__diciembre=False,is_staff=False)

	wb = Workbook()
	#Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
	ws = wb.active
	#En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
	ws['B1'] = 'REPORTE DE INSTRUCTORES %s'%(mes)
	#Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
	ws.merge_cells('B1:E1')
	#Creamos los encabezados desde la celda B3 hasta la E3
	ws['B3'] = 'NOMBRES'
	ws['C3'] = 'APELLIDOS'
	ws['D3'] = 'CEDULA'
	ws['E3'] = 'CORREO' 
	ws['F3'] = 'TELEFONO' 
	ws['G3'] = 'SUPERVISOR'      
	cont=4
	#Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
	for persona in lista_consultar:
		ws.cell(row=cont,column=2).value = persona.first_name
		ws.cell(row=cont,column=3).value = persona.last_name
		ws.cell(row=cont,column=4).value = persona.username
		ws.cell(row=cont,column=5).value = persona.email
		ws.cell(row=cont,column=6).value = persona.user_profile.telefono
		ws.cell(row=cont,column=7).value = persona.user_profile.supervisor
		cont = cont + 1
	#Establecemos el nombre del archivo
	nombre_archivo ="ReporteInstructores%s.xlsx"%(mes)
	#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
	response = HttpResponse(content_type="application/ms-excel") 
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response	

def consultar_sin_subir_view(request,pagina,id_mes):
	meses = { '1':'ENERO', '2':'FEBRERO', '3':'MARZO', '4':'ABRIL', '5':'MAYO','6':'JUNIO','7':'JULIO','8':'AGOSTO','9':'SEPTIEMBRE','10':'OCTUBRE','11':'NOVIEMBRE','12':'DICIEMBRE' }	
	mes = meses[id_mes]	
	num_mes = id_mes
	primera = "<<Primera"
	ultima = "Ultima>>"	
	if id_mes=='1':
		lista_consultar = User.objects.filter(user_profile__enero=False,is_staff=False)
	elif id_mes=='2':
		lista_consultar = User.objects.filter(user_profile__febrero=False,is_staff=False)
	elif id_mes=='3':
		lista_consultar = User.objects.filter(user_profile__marzo=False,is_staff=False)
	elif id_mes=='4':
		lista_consultar = User.objects.filter(user_profile__abril=False,is_staff=False)
	elif id_mes=='5':
		lista_consultar = User.objects.filter(user_profile__mayo=False,is_staff=False)
	elif id_mes=='6':
		lista_consultar = User.objects.filter(user_profile__junio=False,is_staff=False)
	elif id_mes=='7':
		lista_consultar = User.objects.filter(user_profile__julio=False,is_staff=False)
	elif id_mes=='8':
		lista_consultar = User.objects.filter(user_profile__agosto=False,is_staff=False)
	elif id_mes=='9':
		lista_consultar = User.objects.filter(user_profile__septiembre=False,is_staff=False)
	elif id_mes=='10':
		lista_consultar = User.objects.filter(user_profile__octubre=False,is_staff=False)
	elif id_mes=='11':
		lista_consultar = User.objects.filter(user_profile__noviembre=False,is_staff=False)
	elif id_mes=='12':
		lista_consultar = User.objects.filter(user_profile__diciembre=False,is_staff=False)	
	
	usuarios = User.objects.filter(is_staff=False)
	instructores = usuarios.count
	tolal_reportes = lista_consultar.count
	#tolal_reportes = 0
	#for c in consultar_enero:
	#	tolal_reportes=tolal_reportes+1
	query = request.GET.get('q','')     
	if query:
		qset = (
			Q(first_name__icontains=query)|
			Q(last_name__icontains=query)|
			Q(username__icontains=query)|
			Q(email__icontains=query)
		)
		if id_mes=='1':
			results = User.objects.filter(qset,is_staff=False,user_profile__enero=False).distinct()			
		elif id_mes=='2':
			results = User.objects.filter(qset,is_staff=False,user_profile__febrero=False).distinct()			
		elif id_mes=='3':
			results = User.objects.filter(qset,is_staff=False,user_profile__marzo=False).distinct()			
		elif id_mes=='4':
			results = User.objects.filter(qset,is_staff=False,user_profile__abril=False).distinct()			
		elif id_mes=='5':
			results = User.objects.filter(qset,is_staff=False,user_profile__mayo=False).distinct()			
		elif id_mes=='6':
			results = User.objects.filter(qset,is_staff=False,user_profile__junio=False).distinct()			
		elif id_mes=='7':
			results = User.objects.filter(qset,is_staff=False,user_profile__julio=False).distinct()			
		elif id_mes=='8':
			results = User.objects.filter(qset,is_staff=False,user_profile__agosto=False).distinct()			
		elif id_mes=='9':
			results = User.objects.filter(qset,is_staff=False,user_profile__septiembre=False).distinct()			
		elif id_mes=='10':
			results = User.objects.filter(qset,is_staff=False,user_profile__octubre=False).distinct()			
		elif id_mes=='11':
			results = User.objects.filter(qset,is_staff=False,user_profile__noviembre=False).distinct()			
		elif id_mes=='12':
			results = User.objects.filter(qset,is_staff=False,user_profile__diciembre=False).distinct()			

		#results = User.objects.filter(qset,is_staff=False).distinct()  
		mostrar = False      
	else:
		mostrar = True
		results = []

	paginator = Paginator(lista_consultar, 15) 
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		consultar = paginator.page(page)
	except (EmptyPage,InvalidPage):
		consultar = paginator.page(paginator.num_pages)		
		
	return render_to_response('usuarios/consultar_sin_subir.html', {
		"results": results,
		"query": query,
		"mostrar": mostrar,		
		"consultar":consultar,
		"total_reportes":tolal_reportes, 
		"mes":mes,
		"instructores":instructores,
		"num_mes":num_mes,  
		"primera":primera,
		"ultima":ultima,     
	},context_instance=RequestContext(request))


	#ctx = {'consultar':consultar, 'total_reportes':tolal_reportes, 'mes':mes}
	#return render(request, 'usuarios/consultar.html',ctx) 

def consultar_view(request,pagina,id_mes):
	meses = { '1':'ENERO', '2':'FEBRERO', '3':'MARZO', '4':'ABRIL', '5':'MAYO','6':'JUNIO','7':'JULIO','8':'AGOSTO','9':'SEPTIEMBRE','10':'OCTUBRE','11':'NOVIEMBRE','12':'DICIEMBRE' }	
	mes = meses[id_mes]	
	num_mes = id_mes
	primera = "<<Primera"
	ultima = "Ultima>>"	
	lista_consultar = Reporte_Mensual.objects.filter(mes=mes).order_by('fecha')
	usuarios = User.objects.filter(is_staff=False)
	instructores = usuarios.count
	tolal_reportes = lista_consultar.count
	#tolal_reportes = 0
	#for c in consultar_enero:
	#	tolal_reportes=tolal_reportes+1
	query = request.GET.get('q','')     
	if query:
		qset = (
			Q(usuario__first_name__icontains=query)|
			Q(usuario__last_name__icontains=query)|
			Q(usuario__username__icontains=query)|
			Q(usuario__email__icontains=query)
		)
		results = Reporte_Mensual.objects.filter(qset,mes=mes).distinct()  
		mostrar = False      
	else:
		mostrar = True
		results = []

	paginator = Paginator(lista_consultar, 15) 
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		consultar = paginator.page(page)
	except (EmptyPage,InvalidPage):
		consultar = paginator.page(paginator.num_pages)		
		
	return render_to_response('usuarios/consultar.html', {
		"results": results,
		"query": query,
		"mostrar": mostrar,		
		"consultar":consultar,
		"total_reportes":tolal_reportes, 
		"mes":mes,
		"instructores":instructores,
		"num_mes":num_mes,  
		"primera":primera,
		"ultima":ultima,     
	},context_instance=RequestContext(request))
